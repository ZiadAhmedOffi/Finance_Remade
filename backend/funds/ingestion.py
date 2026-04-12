import io
import decimal
from openpyxl import Workbook, load_workbook
from django.db import transaction
from .models import Fund, ModelInput, CurrentDeal, InvestmentDeal, FundLog

class ExcelIngestService:
    """
    Service for generating Excel templates and ingesting fund data.
    """

    @staticmethod
    def generate_template(fund):
        """Generates a downloadable Excel template pre-filled with existing keys."""
        wb = Workbook()
        
        # Sheet 1: Model Inputs
        ws1 = wb.active
        ws1.title = "model_inputs"
        ws1.append(["input_key", "input_value", "data_type"])
        
        # Get all fields of ModelInput except id, fund, and updated_at
        model_input = ModelInput.objects.filter(fund=fund).first()
        if model_input:
            for field in ModelInput._meta.fields:
                if field.name not in ["id", "fund", "updated_at"]:
                    val = getattr(model_input, field.name)
                    ws1.append([field.name, str(val), type(val).__name__])
        
        # Sheet 2: Current Deals
        ws2 = wb.create_sheet("current_deals")
        ws2.append(["company_name", "company_type", "industry", "entry_year", 
                   "latest_valuation_year", "amount_invested", "entry_valuation", 
                   "latest_valuation", "pro_rata_rights"])
        
        # Sheet 3: Future Deals (Prognosis)
        ws3 = wb.create_sheet("future_deals")
        ws3.append(["company_name", "company_type", "industry", "entry_year", 
                   "exit_year", "amount_invested", "entry_valuation", 
                   "base_factor", "downside_factor", "upside_factor", 
                   "selected_scenario", "expected_number_of_rounds"])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def ingest_data(fund, file, actor):
        """Processes the uploaded Excel file and applies changes atomically."""
        try:
            wb = load_workbook(file, data_only=True, read_only=True)
            
            # 1. Validate Sheet Existence
            required_sheets = ["model_inputs", "current_deals", "future_deals"]
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    return False, f"Missing required sheet: {sheet}"

            # 2. Parse & Validate
            errors = []
            
            # A. Model Inputs
            model_updates = {}
            ws_inputs = wb["model_inputs"]
            for row_idx, row in enumerate(ws_inputs.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]: continue
                key, val = row[0], row[1]
                # Type conversion based on field
                try:
                    field = ModelInput._meta.get_field(key)
                    if field.get_internal_type() == 'DecimalField':
                        model_updates[key] = decimal.Decimal(str(val))
                    elif field.get_internal_type() == 'PositiveIntegerField':
                        model_updates[key] = int(val)
                except Exception:
                    errors.append({"sheet": "model_inputs", "row": row_idx, "column": "input_key", "message": f"Invalid key or value: {key}"})

            # B. Current Deals
            current_deals_data = []
            ws_current = wb["current_deals"]
            for row_idx, row in enumerate(ws_current.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]: continue
                try:
                    current_deals_data.append({
                        "fund": fund,
                        "company_name": str(row[0]),
                        "company_type": str(row[1] or ""),
                        "industry": str(row[2] or ""),
                        "entry_year": int(row[3]),
                        "latest_valuation_year": int(row[4]),
                        "amount_invested": decimal.Decimal(str(row[5])),
                        "entry_valuation": decimal.Decimal(str(row[6])),
                        "latest_valuation": decimal.Decimal(str(row[7])),
                        "is_pro_rata": False, # Pro-rata deals cannot be uploaded via Excel
                        "pro_rata_rights": bool(row[8]),
                    })
                except Exception as e:
                    errors.append({"sheet": "current_deals", "row": row_idx, "message": f"Validation error: {str(e)}"})

            # C. Future Deals
            future_deals_data = []
            ws_future = wb["future_deals"]
            for row_idx, row in enumerate(ws_future.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]: continue
                try:
                    future_deals_data.append({
                        "fund": fund,
                        "company_name": str(row[0]),
                        "company_type": str(row[1] or ""),
                        "industry": str(row[2] or ""),
                        "entry_year": int(row[3]),
                        "exit_year": int(row[4]),
                        "amount_invested": decimal.Decimal(str(row[5])),
                        "entry_valuation": decimal.Decimal(str(row[6])),
                        "base_factor": decimal.Decimal(str(row[7] or 1.0)),
                        "downside_factor": decimal.Decimal(str(row[8] or 1.0)),
                        "upside_factor": decimal.Decimal(str(row[9] or 1.0)),
                        "selected_scenario": str(row[10] or "BASE"),
                        "expected_number_of_rounds": int(row[11] or 0),
                        "is_pro_rata": False, # Pro-rata deals cannot be uploaded via Excel
                    })
                except Exception as e:
                    errors.append({"sheet": "future_deals", "row": row_idx, "message": f"Validation error: {str(e)}"})

            if errors:
                return False, errors

            # 3. Apply Atomic Update
            with transaction.atomic():
                # Model Inputs (Overwrite)
                if model_updates:
                    ModelInput.objects.filter(fund=fund).update(**model_updates)
                
                # Deals (Append)
                CurrentDeal.objects.bulk_create([CurrentDeal(**d) for d in current_deals_data])
                InvestmentDeal.objects.bulk_create([InvestmentDeal(**d) for d in future_deals_data])
                
                # Audit Log
                FundLog.objects.create(
                    actor=actor,
                    target_fund=fund,
                    action="MODEL_INPUTS_UPDATED",
                    metadata={
                        "ingestion": "EXCEL_SUCCESS",
                        "model_inputs_updated": len(model_updates),
                        "current_deals_appended": len(current_deals_data),
                        "future_deals_appended": len(future_deals_data)
                    }
                )

            return True, {
                "model_inputs": len(model_updates),
                "current_deals_appended": len(current_deals_data),
                "future_deals_appended": len(future_deals_data)
            }

        except Exception as e:
            return False, f"Unexpected error processing Excel file: {str(e)}"
